from typing import Generator, Tuple
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from src.pacote_log.config_log import logger
from src.services.dados.arquivo import Arquivo
from unidecode import unidecode
import re
import os


class ExcelDados(Arquivo[Workbook]):

    def __init__(self, nome_arquivo: str = None, texto: str = None) -> None:
        super().__init__(nome_arquivo, texto)
        self.__planilha = self._abrir_arquivo()
        self.__nome_aba = self.__planilha.active.title
        self.__aba = self.__planilha[self.__nome_aba]
        self.__ultima_linha = self.__aba.max_row

    def __tratar_texto(self, texto: str) -> str:
        """Método para tratar texto e deixar somente letras

        Args:
            texto (str): recebe um texto

        Returns:
            str: texto tratado
        """
        texto = texto.replace(' ', '_')
        texto = unidecode(texto.lower())
        texto = re.sub('[^a-z_]', '', texto)
        return texto

    def _abrir_arquivo(self) -> Workbook:
        """Método para abrir a planilha

        Returns:
            Workbook: uma planilha
        """
        try:
            planilha = load_workbook(self._caminho_arquivo)
            return planilha
        except FileNotFoundError:
            logger.error('Arquivo não encontrado')
            exit(1)
        except Exception as e:
            logger.critical(f'FALHA TOTAL: {e}')

    def ler_valores(self) -> Generator[Tuple[str], None, None]:
        """Método para ler os dados de arquivo, banco

        Yields:
            Generator[Tuple[str, str], None, None]: Gerador que retorna a url e o nome do vídeo
        """
        for linha in self.__aba.iter_rows(min_row=2, max_col=3):
            url, nome_video, marcador = linha[:3]

            if marcador.value != 'X':
                try:
                    print('Fazendo resumo', url.value,
                          nome_video.value, marcador.value)
                    yield url.value.split('=')[-1], self.__tratar_texto(nome_video.value)

                except:
                    break

    def gravar_dados(self,):
        try:
            ws = self.__planilha.active
            for linha in range(1, self.__ultima_linha):
                celula = ws.cell(row=linha, column=3)
                if celula.value is None or celula.value == '':
                    celula.value = 'X'

        except Exception as e:
            logger.critical(f'ERRO FATAL: {e}')
            exit(1)

    def salvar_dados(self, nome_arquivo: str = None, **kwargs):
        try:
            self.gravar_dados()
            self.__planilha.save(self._caminho_arquivo)
        except OSError as e:
            logger.error(f'Erro de sistema ao salvar o arquivo: {e}')
            exit(1)
        except Exception as e:
            logger.error(f'Erro ao salvar o arquivo: {e}')
            exit(1)

    def __del__(self):
        self.__planilha.close()
